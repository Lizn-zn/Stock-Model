# -*- coding: utf-8 -*-
"""
Created on Tue May 13 16:00:45 2025

@author: FW297EG
"""
from openpyxl import load_workbook
import re
# 加载 Excel 文件
wb = load_workbook(filename=r'C:\Users\FW297EG\OneDrive - EY CHINA\Desktop\数据项目\测试\IT-致欧科技24-MA-NEYA系统.TOC01-310账号新增.xlsx')

# 选择指定的 Sheet
sheet_name = 'IT一般控制测试'
ws = wb[sheet_name]

# 读取指定单元格（例如 B3）
cell_value = ws['Z59'].value
if cell_value is not None and str(cell_value).strip() != '':
    print(f"单元格 Z59非空，内容为: {cell_value}")
    pattern = r'<IPE[^>]*>'
    matches = re.findall(pattern, cell_value)
    #gc里的所有的IPE名称
    ipe_names = [match[1:-1] for match in matches] 

    sheet_names = wb.sheetnames
    #excel里所有ipe开头的sheet_name
    filtered_sheets = [name for name in sheet_names if name.startswith("IPE")]

    matched_sheets = [sheet for sheet in ipe_names if sheet not in filtered_sheets]

    print("Not Matched Sheets:", matched_sheets)
else:
    print(f"单元格 Z59 是空的")
